import binascii
import json
import pty
import os
import argparse
import signal
# xlsx excel file handling
#import pandas as pd
from openpyxl import load_workbook
import serial
import serial.rs485
import sys
import pymodbus
import logging
from pymodbus.version import version
from pymodbus.server.sync import StartSerialServer
from pymodbus.device import ModbusDeviceIdentification
from pymodbus.datastore import ModbusSequentialDataBlock, ModbusSparseDataBlock
from pymodbus.datastore import ModbusSlaveContext, ModbusServerContext
from pymodbus.transaction import ModbusRtuFramer, ModbusBinaryFramer

log = None

# Create child class to be able to use PTY, master /dev/ptmx requires use of filedescriptor
class ModbusSerialServerPTY(pymodbus.server.sync.ModbusSerialServer):
    #override methods
    def __init__(self, context, framer=None, identity=None, fd=None, **kwargs):
        self.fd = fd
        super().__init__(context, framer, identity, **kwargs)
        #self.fd = kwargs.get('fd',  0)  # filedescriptor


    def _connect(self):
        super()._connect()
        if self.socket is not None:
            self.socket.fd = self.fd
        return self.socket is not None


class ModbusRtuEmulator:
    pty_master_fd = None
    pty_slave_fd = None

    def __init__(self, master_ttyname, master_fd, slave_fd):
        self.log = None
        self.pty_master_dev = master_ttyname
        ModbusRtuEmulator.pty_master_fd = master_fd
        ModbusRtuEmulator.pty_slave_fd = slave_fd

    def config_logging(self):
        # --------------------------------------------------------------------------- #
        # configure the service logging
        # --------------------------------------------------------------------------- #
        FORMAT = ('%(asctime)-15s %(threadName)-15s'
                  ' %(levelname)-8s %(module)-15s:%(lineno)-8s %(message)s')
        logging.basicConfig(format=FORMAT)
        global log
        log = logging.getLogger()
        if args.debug:
            log.setLevel(logging.DEBUG)
        else:
            log.setLevel(logging.INFO)
        self.log = log

    def CreateDataBlocksFromXLSX(self):
        # sub function to
        def create_register_values(ws):
            registers = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                pass
                address = int(row[0].split(':')[1])
                minval = int(row[3])
                if type(row[4]) == str:
                    maxval = int(row[4], 0)
                else:
                    maxval = row[4]
                initval = int(maxval - ((maxval - minval) // 2))
                registers[address] = initval
            return registers
        datablocks = {}
        # Load the xlsx file
        #excel_data = pd.read_excel('ICS.2 Modbus Registers 2019-01-10.xlsx')
        # Read the values of the file in the dataframe
        #data = pd.DataFrame(excel_data, columns=['Sales Date', 'Sales Person', 'Amount'])
        wb = load_workbook(filename=os.path.dirname(__file__) +'/ICS.2 Modbus Registers 2019-01-10.xlsx')
        datablocks['co'] = create_register_values(wb['Coils'])
        datablocks['di'] = create_register_values(wb['Inputs']) # discrete inputs
        datablocks['ir'] = create_register_values(wb['Input Registers'])
        datablocks['hr'] = create_register_values(wb['Holding Registers'])
        return datablocks

    def run_server(self):
        # setup registers
        # create datablocks from excel file
        datablocks = self.CreateDataBlocksFromXLSX()

        slavecontext = ModbusSlaveContext(
            di=ModbusSparseDataBlock(datablocks['di'], False),
            co=ModbusSparseDataBlock(datablocks['co'], False),
            hr=ModbusSparseDataBlock(datablocks['hr'], False),
            ir=ModbusSparseDataBlock(datablocks['ir'], False))
            #di=ModbusSequentialDataBlock(0, [1]*11320),
            #co=ModbusSequentialDataBlock(0, [1]*11256),
            #hr=ModbusSequentialDataBlock(0, [257]*14648),
            #ir=ModbusSequentialDataBlock(0, [257]*11640))
        # setup slave id's
        slaves = {
                 0x01: slavecontext
        }
        # setup the context
        context = ModbusServerContext(slaves=slaves, single=False)
        # ----------------------------------------------------------------------- #
        # initialize the server information
        # ----------------------------------------------------------------------- #
        identity = ModbusDeviceIdentification()
        identity.VendorName = 'Pymodbus'
        identity.ProductCode = 'PM'
        identity.VendorUrl = 'http://github.com/riptideio/pymodbus/'
        identity.ProductName = 'Pymodbus Server'
        identity.ModelName = 'Pymodbus Server'
        identity.MajorMinorRevision = version.short()
        # Run the server
        #StartSerialServer(context, framer=ModbusRtuFramer, identity=identity, port=master_ttyname,
        #                  timeout=.005, baudrate=38400, fd=master_fd)
        server = ModbusSerialServerPTY(context, framer=ModbusRtuFramer, identity=identity,
                                       port=self.pty_master_dev, timeout=.005, baudrate=38400, fd=self.pty_master_fd)
        server.serve_forever()


def run_master_as_raw_file(master):
    # setup raw master connection
    while True:
        req = os.read(master, 1)
        try:
            req_str_ascii = str(req, 'UTF-8')
        except UnicodeDecodeError:
            req_str_ascii = ""
        print("<-- RX: 0x" + req.hex() + "\t\t/\tASCII: " + req_str_ascii)


def run_master_as_serial(master_ttyname, master_fd):
    # setup master connection as serial port
    master_ser = serial.rs485.RS485(master_ttyname, 38400)
    # since filename of master file descriptor is always /dev/ptmx assign the real fd
    master_ser.fd = master_fd
    # just get and print all serial data
    while True:
        req = master_ser.read(1)
        try:
            req_str_ascii = str(req, 'UTF-8')
        except UnicodeDecodeError:
            req_str_ascii = ""
        print("<-- RX: 0x" + req.hex() + "\t\t/\tASCII: " + req_str_ascii)


def update_configfile(slave_ttyname):
    conf_file_path = os.path.dirname(__file__) + '/configuration.json' #  realpath
    if not os.path.isfile(conf_file_path):
        print('ERROR: \"' + conf_file_path + '\" does not exist!!!\n')
        return
    # File exist
    with open(conf_file_path, 'r+') as f:
        json_data = json.load(f)
        if json_data['device'] == slave_ttyname:
            print('INFO: ' + conf_file_path + ' already set to: ' + slave_ttyname + ' , nothing to do.\n')
            return
        # write new device to json file
        json_data['device'] = slave_ttyname
        f.seek(0)
        f.write(json.dumps(json_data, indent=4) + "\n")
        f.truncate()
        print('Successfuly updated config file: ' + conf_file_path + ' to use serial terminal device: ' + slave_ttyname + '\n')


def main():
    #process_parent_child()
    pty_master_fd, pty_slave_fd = pty.openpty()
    master_ttyname = os.ttyname(pty_master_fd)
    slave_ttyname = os.ttyname(pty_slave_fd)

    print('filename of the Master: ' + master_ttyname)
    print('filedescriptor of master: ' + str(pty_master_fd))
    print('Filename of the Slave: ' + slave_ttyname)
    # Update configuration.json used by nodered to use the new slave pseudo terminal, if doing so is requested by input argument
    if args.updateconfig:
        update_configfile(slave_ttyname)
    #run_master_as_raw_file(master)
    #run_master_as_serial(master_ttyname, pty_master_fd)
    # start the modbus server
    modbus_rtu_emu = ModbusRtuEmulator(master_ttyname, pty_master_fd, pty_slave_fd)
    modbus_rtu_emu.config_logging()
    modbus_rtu_emu.run_server()



if __name__ == '__main__':
    print('Running ModbusRtuEmulator.py...\n')
    argparser = argparse.ArgumentParser()
    argparser.add_argument("-u", "--updateconfig", help="Update configuration.json file with the pseudo terminal "
                                                        "serial slave created", action="store_true")
    argparser.add_argument("-d", "--debug", help="Set log level to debug", action="store_true")
    args = argparser.parse_args()

    try:
        main()
    except KeyboardInterrupt:
        print('Interrupted')
        try:
            os.close(ModbusRtuEmulator.pty_slave_fd)
            os.close(pty_master_fd)
            sys.exit(0)
        except SystemExit:
            os._exit(0)
